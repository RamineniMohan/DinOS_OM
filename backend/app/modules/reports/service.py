import csv
import io
import uuid
from decimal import Decimal

from sqlalchemy import Date, cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.billing.models import Invoice
from app.modules.inventory.models import Ingredient, Unit
from app.modules.orders.models import Order, OrderStatus
from app.modules.reports.schemas import (
    DateRangeParams,
    GSTReport,
    GSTReportRow,
    InventoryReport,
    InventoryReportRow,
    SalesReport,
    SalesReportRow,
)


class ReportsService:

    # ── Sales ──────────────────────────────────────────────────────────────
    @staticmethod
    async def sales_report(db: AsyncSession, params: DateRangeParams) -> SalesReport:
        result = await db.execute(
            select(
                cast(Order.created_at, Date).label('date'),
                func.count(Order.id).label('order_count'),
                func.sum(Order.total_amount).label('total_revenue'),
                func.sum(Order.tax_amount).label('total_tax'),
            )
            .where(
                Order.restaurant_id == params.restaurant_id,
                cast(Order.created_at, Date) >= params.start_date,
                cast(Order.created_at, Date) <= params.end_date,
                Order.status != OrderStatus.CANCELLED,
            )
            .group_by(cast(Order.created_at, Date))
            .order_by(cast(Order.created_at, Date))
        )
        rows_raw = result.all()
        report_rows = [
            SalesReportRow(
                date=r.date,
                order_count=r.order_count,
                total_revenue=r.total_revenue or Decimal('0'),
                total_tax=r.total_tax or Decimal('0'),
                average_order_value=(r.total_revenue / r.order_count) if r.order_count else Decimal('0'),
            )
            for r in rows_raw
        ]
        return SalesReport(
            restaurant_id=params.restaurant_id,
            start_date=params.start_date,
            end_date=params.end_date,
            total_orders=sum(r.order_count for r in report_rows),
            total_revenue=sum(r.total_revenue for r in report_rows),
            total_tax=sum(r.total_tax for r in report_rows),
            rows=report_rows,
        )

    # ── GST ────────────────────────────────────────────────────────────────
    @staticmethod
    async def gst_report(db: AsyncSession, params: DateRangeParams) -> GSTReport:
        result = await db.execute(
            select(Invoice)
            .where(
                Invoice.restaurant_id == params.restaurant_id,
                cast(Invoice.created_at, Date) >= params.start_date,
                cast(Invoice.created_at, Date) <= params.end_date,
            )
            .order_by(Invoice.created_at)
        )
        invoices = result.scalars().all()
        rows = [
            GSTReportRow(
                invoice_number=inv.invoice_number,
                date=inv.created_at,
                customer_name=inv.customer_name,
                customer_phone=inv.customer_phone,
                customer_gstin=inv.customer_gstin,
                subtotal=inv.subtotal,
                cgst=inv.cgst_amount,
                sgst=inv.sgst_amount,
                igst=inv.igst_amount,
                total_tax=inv.total_tax,
                total_amount=inv.total_amount,
            )
            for inv in invoices
        ]
        return GSTReport(
            restaurant_id=params.restaurant_id,
            start_date=params.start_date,
            end_date=params.end_date,
            rows=rows,
        )

    # ── Inventory ──────────────────────────────────────────────────────────
    @staticmethod
    async def inventory_report(db: AsyncSession, restaurant_id: uuid.UUID) -> InventoryReport:
        result = await db.execute(
            select(Ingredient, Unit)
            .join(Unit, Ingredient.unit_id == Unit.id)
            .where(Ingredient.restaurant_id == restaurant_id)
            .order_by(Ingredient.name)
        )
        rows = [
            InventoryReportRow(
                ingredient_name=ing.name,
                unit=unit.abbreviation,
                current_stock=ing.current_stock,
                low_stock_threshold=ing.low_stock_threshold,
                is_low_stock=ing.current_stock <= ing.low_stock_threshold,
            )
            for ing, unit in result.all()
        ]
        return InventoryReport(restaurant_id=restaurant_id, rows=rows)

    # ── CSV exports ────────────────────────────────────────────────────────
    @staticmethod
    async def sales_report_csv(db: AsyncSession, params: DateRangeParams) -> str:
        report = await ReportsService.sales_report(db, params)
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(['Date', 'Orders', 'Revenue', 'Tax', 'Avg Order Value'])
        for r in report.rows:
            w.writerow([r.date, r.order_count, r.total_revenue, r.total_tax, r.average_order_value])
        return out.getvalue()

    @staticmethod
    async def gst_report_csv(db: AsyncSession, params: DateRangeParams) -> str:
        report = await ReportsService.gst_report(db, params)
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(['Invoice No', 'Date', 'Customer', 'GSTIN', 'Subtotal', 'CGST', 'SGST', 'IGST', 'Total Tax', 'Total'])
        for r in report.rows:
            w.writerow([
                r.invoice_number,
                r.date.strftime('%Y-%m-%d %H:%M') if r.date else '',
                r.customer_name or '', r.customer_gstin or '',
                r.subtotal, r.cgst, r.sgst, r.igst, r.total_tax, r.total_amount,
            ])
        return out.getvalue()

    @staticmethod
    async def inventory_report_csv(db: AsyncSession, restaurant_id: uuid.UUID) -> str:
        report = await ReportsService.inventory_report(db, restaurant_id)
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(['Ingredient', 'Unit', 'Current Stock', 'Low Stock Threshold', 'Low Stock?'])
        for r in report.rows:
            w.writerow([r.ingredient_name, r.unit, r.current_stock, r.low_stock_threshold, r.is_low_stock])
        return out.getvalue()

    # ── XLSX exports ───────────────────────────────────────────────────────
    @staticmethod
    async def sales_report_xlsx(db: AsyncSession, params: DateRangeParams) -> bytes:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        report = await ReportsService.sales_report(db, params)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sales Report'
        headers = ['Date', 'Orders', 'Revenue (INR)', 'Tax (INR)', 'Avg Order Value (INR)']
        header_fill = PatternFill('solid', fgColor='1E3A5F')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row_idx, r in enumerate(report.rows, 2):
            ws.append([str(r.date), r.order_count, float(r.total_revenue), float(r.total_tax), float(r.average_order_value)])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    async def gst_report_xlsx(db: AsyncSession, params: DateRangeParams) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        report = await ReportsService.gst_report(db, params)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'GST Report'
        headers = ['Invoice No', 'Date', 'Customer', 'GSTIN', 'Subtotal', 'CGST', 'SGST', 'IGST', 'Total Tax', 'Total']
        fill = PatternFill('solid', fgColor='1E3A5F')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = fill
        for r in report.rows:
            ws.append([
                r.invoice_number,
                r.date.strftime('%Y-%m-%d %H:%M') if r.date else '',
                r.customer_name or '', r.customer_gstin or '',
                float(r.subtotal), float(r.cgst), float(r.sgst), float(r.igst), float(r.total_tax), float(r.total_amount),
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── PDF exports ────────────────────────────────────────────────────────
    @staticmethod
    async def sales_report_pdf(db: AsyncSession, params: DateRangeParams) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        report = await ReportsService.sales_report(db, params)
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(f'Sales Report: {params.start_date} to {params.end_date}', styles['Title']),
            Spacer(1, 12),
        ]
        data = [['Date', 'Orders', 'Revenue (INR)', 'Tax (INR)', 'AOV (INR)']]
        for r in report.rows:
            data.append([str(r.date), r.order_count, f'{r.total_revenue:.2f}', f'{r.total_tax:.2f}', f'{r.average_order_value:.2f}'])
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ]))
        elements.append(t)
        doc.build(elements)
        return buf.getvalue()

    @staticmethod
    async def gst_report_pdf(db: AsyncSession, params: DateRangeParams) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        report = await ReportsService.gst_report(db, params)
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(f'GST Report (GSTR Reference): {params.start_date} to {params.end_date}', styles['Title']),
            Spacer(1, 12),
        ]
        data = [['Invoice No', 'Date', 'Customer', 'GSTIN', 'Subtotal', 'CGST', 'SGST', 'IGST', 'Tax', 'Total']]
        for r in report.rows:
            data.append([
                r.invoice_number,
                r.date.strftime('%Y-%m-%d') if r.date else '',
                (r.customer_name or '')[:20], r.customer_gstin or '',
                f'{r.subtotal:.2f}', f'{r.cgst:.2f}', f'{r.sgst:.2f}', f'{r.igst:.2f}', f'{r.total_tax:.2f}', f'{r.total_amount:.2f}',
            ])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ]))
        elements.append(t)
        doc.build(elements)
        return buf.getvalue()
